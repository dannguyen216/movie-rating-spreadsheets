import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors
import sys


# Global variables for Worksheet styles
# Cell colors
LIGHT_RED_FILL = PatternFill(fill_type='solid',
                             start_color='FF9999',
                             end_color='FF9999')
LIGHT_BLUE_FILL = PatternFill(fill_type='solid',
                              start_color='CCCCFF',
                              end_color='CCCCFF')
RED_FILL = PatternFill(fill_type='solid',
                       start_color='FF6666',
                       end_color='FF6666')
ORANGE_FILL = PatternFill(fill_type='solid',
                          start_color='FFB266',
                          end_color='FFB266')
YELLOW_FILL = PatternFill(fill_type='solid',
                          start_color='FFFF66',
                          end_color='FFFF66')
GREEN_FILL = PatternFill(fill_type='solid',
                         start_color='66FF66',
                         end_color='66FF66')
BLACK_FILL = PatternFill(fill_type='solid',
                         start_color='000000',
                         end_color='000000')

# Cell value alignment
CENTER_ALIGN = Alignment(horizontal='center')

# Font Styles
COLUMN_TITLE_FONT = Font(size=20, bold=True, underline='single')
BOLD_FONT = Font(bold=True)
AVERAGE_FONT = Font(bold=True, color=colors.WHITE)

# Border Styles
CELL_BORDER = Border(left=Side(border_style='thick'),
                     right=Side(border_style='thick'),
                     top=Side(border_style='thick'),
                     bottom=Side(border_style='thick'))


# Function that reads an input file given by the user in a specific
# specific format and returns a list of movies along with their
# corresponding rating and release date.
# Each element in the list will contain information in the following
# format:
# [Movie Title, Movie Rating, Release Date]
def read_movie_input(input_file):
    movie_data_list = []
    # Open the file and read it line by line
    with open(input_file, 'r') as file:
        for line in file.readlines():
            # Only append movie data if the line if the line has characters
            # other than whitespaces.
            if line.strip():
                # Remove trailing whitespaces in the values and read
                # in elements separated by three semicolons
                movie_data = [s.strip() for s in line.split(';;;')]
                # Valid data entries should have 3 elements in the list
                # after parsing the line.
                if len(movie_data) == 3:
                    movie_data_list.append(movie_data)

    return movie_data_list


# Function used to write the titles for each column on the worksheet.
# The three columns are:
#    A.) Movie Title
#    B.) Rating (A review score for the movie from 1 to 5)
#    C.) Release Date of the movie
def write_column_names(worksheet):
    # Label the Movie Title column and edit the cell styles
    worksheet['A1'] = "Movie Title"
    worksheet['A1'].font = COLUMN_TITLE_FONT
    worksheet['A1'].alignment = CENTER_ALIGN
    worksheet['A1'].fill = LIGHT_RED_FILL
    worksheet['A1'].border = CELL_BORDER
    worksheet.column_dimensions['A'].width = 50

    # Label the Movie Rating column and edit the cell styles
    worksheet['B1'] = "Rating"
    worksheet['B1'].font = COLUMN_TITLE_FONT
    worksheet['B1'].alignment = CENTER_ALIGN
    worksheet['B1'].fill = LIGHT_RED_FILL
    worksheet['B1'].border = CELL_BORDER
    worksheet.column_dimensions['B'].width = 25

    # label the Release Date column and edit the cell styles
    worksheet['C1'] = "Release Date"
    worksheet['C1'].font = COLUMN_TITLE_FONT
    worksheet['C1'].alignment = CENTER_ALIGN
    worksheet['C1'].fill = LIGHT_RED_FILL
    worksheet['C1'].border = CELL_BORDER
    worksheet.column_dimensions['C'].width = 25

    return


# Function that takes the movie data read from the input file
# and writes the data to the excel spreadsheet given as
# a parameter.
def write_movie_data_to_spreadsheet(worksheet, movie_data_list):
    # The movie data begins on the second row of the spreadsheet
    # (The first row contains the column titles)
    row_num = 2

    # Iterate through each movie (and corresponding information)
    # obtained from the input file
    for movie_data in movie_data_list:
        # Each movie title will be written in the 'A' column
        # in separate rows.
        movie_title = movie_data[0]
        movie_title_cell = 'A' + str(row_num)
        worksheet[movie_title_cell] = movie_title

        # Movie Title cell style edits
        worksheet[movie_title_cell].font = BOLD_FONT
        worksheet[movie_title_cell].alignment = CENTER_ALIGN
        worksheet[movie_title_cell].fill = LIGHT_BLUE_FILL
        worksheet[movie_title_cell].border = CELL_BORDER

        # Movie Ratings will be written to the 'B' column
        rating = movie_data[1]
        rating_cell = 'B' + str(row_num)
        # On the spreadsheet, the rating will be written
        # to the cell along with ' / 5' to signify
        # the intended maximum score.
        worksheet[rating_cell] = '%s / 5' % rating

        # Rating cell style edits
        worksheet[rating_cell].font = BOLD_FONT
        worksheet[rating_cell].alignment = CENTER_ALIGN
        # In order to determine the color of the Movie Rating cell,
        # a helper function get_rating_color is called.
        worksheet[rating_cell].fill = get_rating_color(float(rating))
        worksheet[rating_cell].border = CELL_BORDER

        # Release Dates will be written to the 'C' column
        release_date = movie_data[2]
        date_cell = 'C' + str(row_num)
        worksheet[date_cell] = release_date

        # Release Date cell style edits
        worksheet[date_cell].font = BOLD_FONT
        worksheet[date_cell].alignment = CENTER_ALIGN
        worksheet[date_cell].fill = LIGHT_BLUE_FILL
        worksheet[date_cell].border = CELL_BORDER

        row_num += 1

    return


# A helper function to the write_movie_data_to_spreadsheet
# function that determines the color of each cell in the
# ratings column, depending on what range the rating
# falls under. The ratings are as follows:
#    rating < 2: Red cell color (Bad movie)
#    2 < rating < 3: Orange cell color (Mediocre Movie)
#    rating = 3: Yellow cell color (Decent Movie)
#    rating > 3: Green cell color (Good Movie)
def get_rating_color(rating):
    if rating < 2:
        return RED_FILL
    elif rating < 3:
        return ORANGE_FILL
    elif rating == 3:
        return YELLOW_FILL
    else:
        return GREEN_FILL


# Function that is called after writing data to the worksheet
# After reading and parsing the movie ratings, the function writes
# the average rating to the cell below the last row of the table.
# The average is rounded to the hundreths place (2 decmial places)
def write_average_to_spreadsheet(worksheet):
    # The LEFT and SEARCH functions are used to parse the text in the
    # 'B' column. The text in the Movie Ratings column is listed
    # out of a maximum score of 5, and the Excel text functions
    # are used to isolate the actual number score from the input file.
    rating_string = 'LEFT(B2, SEARCH(\"/\", B2) - 1)'
    for cell in worksheet['B'][2:]:
        # Each cell's value will be obtained from the Excel text
        # funcitons and separated by commas.
        rating_string += ','
        cell_string = cell.column + str(cell.row)
        rating_string += 'LEFT({}, SEARCH(\"/\", {}) - 1)'.format(cell_string,
                                                                  cell_string)

    # The row that the average will be written to is immediately
    # below the last row that the movie data was written to.
    average_row = str(worksheet.max_row + 1)

    # The Average rating will be labeled in the 'A' column
    # next to the average value
    label_cell = 'A' + average_row
    worksheet[label_cell] = 'AVERAGE RATING'
    worksheet[label_cell].font = AVERAGE_FONT
    worksheet[label_cell].alignment = CENTER_ALIGN
    worksheet[label_cell].fill = BLACK_FILL

    average_cell = 'B' + average_row

    # The average will be calculated from the values read from the cells
    # and rounded to the hundreths place.
    worksheet[average_cell] = '=ROUND(AVERAGE({}), 2)'.format(rating_string)
    worksheet[average_cell].font = AVERAGE_FONT
    worksheet[average_cell].alignment = CENTER_ALIGN
    worksheet[average_cell].fill = BLACK_FILL
    return


def main():
    # Currently, the program only accepts a single command line argument
    # in addition to movie_rating_sheets.py
    # If the number of command line arguments is not correct, the
    # correct usage format will be printed, and the program will
    # stop running.
    if len(sys.argv) != 2:
        print('Usage: python movie_rating_sheets.py input_file')
        return 1

    # The input file will be read from the command line
    input_file = sys.argv[1]

    # The file name will exclude the file format to title the
    # spreadsheet that will be created
    file_name = input_file.split('.')[0]

    # The movie information gained from the input file will be
    # stored in a list
    try:
        movie_data_list = read_movie_input(input_file)

    # If the input file cannot be opened in the current directory,
    # notify the user and exit.
    except FileNotFoundError:
        print('Input file \"%s\" not found in current directory.' % input_file)
        print('Exiting...')
        return 1

    # The spreadsheet will only be written if data was able to be obtained
    # from the input file
    if len(movie_data_list) > 0:
        try:
            # Attempt to load the workbook
            print('Loading Movie_Ratings workbook...')
            workbook = openpyxl.load_workbook('Movie_Ratings.xlsx')
            # If a worksheet with the same name as the input file
            # exists, the file will not be overwritten and will
            # exit instead
            if file_name in workbook.sheetnames:
                print('Spreadsheet of name \"%s\" already in workbook.' %
                      file_name)
                print('Exiting...')
                return 1
            else:
                # Create a sheet with the same name as the input file
                workbook.create_sheet(title=file_name)

        except FileNotFoundError:
            # If Movie_Ratings.xlsx is not found in the current directory,
            # the file will be created with a worksheet named after the input
            # file
            print('\nMovie ratings workbook not found in current directory.')
            print('Creating workbook file Movie_Ratings.xlsx...')
            workbook = openpyxl.Workbook()
            workbook.active.title = file_name

        # Create a worksheet with the same name as the input file
        # and write the column titles to the worksheet
        worksheet = workbook.get_sheet_by_name(file_name)
        write_column_names(worksheet)
        print('\nWriting data to worksheet %s...' % worksheet.title)

        # With the data from movie_data_list, write the movie information
        # to the spreadsheet on separate rows
        write_movie_data_to_spreadsheet(worksheet, movie_data_list)

        # Calculate the average rating and write the value to the cell
        # below the movie data
        write_average_to_spreadsheet(worksheet)

        # Attempt to save the changes to the workbook
        print('Saving changes to workbook...')
        try:
            workbook.save('Movie_Ratings.xlsx')
            print('\nChanges saved successfully.')

        except OSError:
            # If the file is still open, changes cannot be saved.
            # The user is notified, and the program exits.
            print('\nFailed to save changes.')
            print('Workbook Movie_Ratings.xlsx is busy.')
            print('Confirm that the workbook is closed before running' +
                  ' the script.')
            print('Exiting...')
            return 1

        return 0

    # If none of the expected information was able to be obtained
    # from the input file, notify the user and exit.
    else:
        print('No movie data obtained from input file %s' % input_file)
        print('Exiting...')
        return 1


if __name__ == '__main__':
    main()
